"""
Excel binary file parser (.xlsx and .xls).
Extracts pure cell data using openpyxl (XLSX) or xlrd (XLS),
producing the same output structure as xml_parser.py.
"""
import io
import os
import time

import openpyxl
import xlrd
from xml_parser import col_to_letter, build_headers


def _cell_to_str(value) -> str:
    """Convert a cell value to a stripped string, matching xml_parser behaviour."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def _is_xls(source) -> bool:
    """Detect old .xls format by path extension or magic bytes."""
    if isinstance(source, (str, os.PathLike)):
        return str(source).lower().endswith(".xls") and not str(source).lower().endswith(".xlsx")
    if isinstance(source, bytes):
        return source[:8].startswith(b"\xd0\xcf\x11\xe0")
    return False


def _parse_xls(source, header_row: int = 1) -> dict:
    """Parse an XLS (BIFF) workbook using xlrd."""
    if isinstance(source, bytes):
        wb = xlrd.open_workbook(file_contents=source)
    else:
        wb = xlrd.open_workbook(source)

    result = {"sheets": {}}

    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        rows_data = []
        max_col = 0

        for row_idx in range(ws.nrows):
            cells = {}
            for col_idx in range(ws.ncols):
                value = _cell_to_str(ws.cell_value(row_idx, col_idx))
                if value:
                    col_key = col_to_letter(col_idx + 1)
                    cells[col_key] = value
                    if col_idx + 1 > max_col:
                        max_col = col_idx + 1

            if cells:
                rows_data.append({"_row": row_idx + 1, "cells": cells})

        headers, header_comments = build_headers(rows_data, max_col, header_row)

        result["sheets"][sheet_name] = {
            "headers": headers,
            "header_comments": header_comments,
            "rows": rows_data,
            "row_count": len(rows_data),
            "col_count": max_col,
            "header_row": header_row,
        }

    return result


def _parse_xlsx(source, header_row: int = 1) -> dict:
    """Parse an XLSX (Office Open XML) workbook using openpyxl."""
    if isinstance(source, (str, os.PathLike)) and os.path.isfile(source):
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    elif isinstance(source, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(source), read_only=True, data_only=True)
    elif hasattr(source, "read"):
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    else:
        raise ValueError(f"Unsupported source type: {type(source)}")

    result = {"sheets": {}}

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data = []
            max_col = 0

            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                cells = {}
                for col_idx, cell in enumerate(row, start=1):
                    value = _cell_to_str(cell.value)
                    if value:
                        col_key = col_to_letter(col_idx)
                        cells[col_key] = value
                        if col_idx > max_col:
                            max_col = col_idx

                if cells:
                    rows_data.append({"_row": row_idx, "cells": cells})

            headers, header_comments = build_headers(rows_data, max_col, header_row)

            result["sheets"][sheet_name] = {
                "headers": headers,
                "header_comments": header_comments,
                "rows": rows_data,
                "row_count": len(rows_data),
                "col_count": max_col,
                "header_row": header_row,
            }
    finally:
        wb.close()

    return result


def parse_workbook(source, header_row: int = 1) -> dict:
    """
    Parse an Excel workbook (.xlsx or .xls) from a file path, bytes, or file-like object.

    Returns the same structure as xml_parser.parse_workbook.
    """
    t0 = time.perf_counter()

    if _is_xls(source):
        result = _parse_xls(source, header_row=header_row)
    else:
        result = _parse_xlsx(source, header_row=header_row)

    result["_parse_ms"] = round((time.perf_counter() - t0) * 1000, 1)
    return result


def parse_file(filepath: str, header_row: int = 1) -> dict:
    """Convenience wrapper: parse a file and include file metadata."""
    data = parse_workbook(filepath, header_row=header_row)
    data["file"] = os.path.basename(filepath)
    data["file_path"] = filepath
    data["file_size"] = os.path.getsize(filepath)
    return data


def parse_bytes(content: bytes, header_row: int = 1) -> dict:
    """Parse Excel content from raw bytes (e.g. from svn cat)."""
    return parse_workbook(content, header_row=header_row)
